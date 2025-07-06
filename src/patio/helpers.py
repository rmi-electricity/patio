from __future__ import annotations

import datetime
import logging
import os
import subprocess
import time
from collections.abc import Callable  # noqa: TC003
from functools import lru_cache
from pathlib import Path  # noqa: TC003

import cvxpy as cp
import numpy as np
import pandas as pd
import polars as pl
from etoolbox.utils.cloud import get
from etoolbox.utils.misc import all_logging_disabled
from numba import njit
from plotly import graph_objs as go  # noqa: TC002
from plotly import io as pio

from patio.constants import RE_TECH, ROOT_PATH, STATE_BOUNDS

LOGGER = logging.getLogger("patio")
__all__ = [
    "add_states",
    "adjust_col_for_pct_owned",
    "agg_profile",
    "cat_multiindex_as_col",
    "check_lat_lons",
    "clean_col",
    "combine_cols",
    "combine_profiles",
    "df_in_half",
    "df_product",
    "distance",
    "distance_arrays",
    "find_na",
    "fix_cems_datetime",
    "fix_na_neg_col",
    "idfn",
    "isclose",
    "load_or_rebuild_parquet",
    "lstrip0",
    "mo_to_days",
    "ninja_profile_fix",
    "prep_for_re_ninja",
    "round_coordinates",
    "seconds_since_touch",
    "show",
]


def df_query[T: pd.DataFrame | pl.DataFrame | pl.LazyFrame](df: T, qs: dict) -> T:
    if isinstance(df, pd.DataFrame):
        for k, v in qs.items():
            if not isinstance(v, dict) or "comp" not in v:
                continue
            if v.get("alt_name", "") in df.columns:
                k = v["alt_name"]
            if v["comp"] == "in":
                df = df[df[k].isin(v["item"])]
            if v["comp"] == "eq":
                df = df[df[k] == v["item"]]
            if v["comp"] == "ge":
                df = df[df[k] >= v["item"]]
            if v["comp"] == "le":
                df = df[df[k] <= v["item"]]
            if v["comp"] == "ge.dt.year":
                df = df[df[k].dt.year >= v["item"]]
            if v["comp"] == "le.dt.year":
                df = df[df[k].dt.year <= v["item"]]
            if v["comp"] == "is_true":
                df = df[df[k]]
            if v["comp"] == "is_false":
                df = df[~df[k]]
        return df
    elif isinstance(df, pl.DataFrame | pl.LazyFrame):
        for k, v in qs.items():
            if not isinstance(v, dict) or "comp" not in v:
                continue
            if v.get("alt_name", "") in df.columns:
                k = v["alt_name"]
            if v["comp"] == "in":
                df = df.filter(pl.col(k).is_in(v["item"]))
            if v["comp"] == "eq":
                df = df.filter(pl.col(k) == v["item"])
            if v["comp"] == "ge":
                df = df.filter(pl.col(k) >= v["item"])
            if v["comp"] == "le":
                df = df.filter(pl.col(k) <= v["item"])
            if v["comp"] == "ge.dt.year":
                df = df.filter(pl.col(k).dt.year() >= v["item"])
            if v["comp"] == "le.dt.year":
                df = df.filter(pl.col(k).dt.year() <= v["item"])
            if v["comp"] == "is_true":
                df = df.filter(pl.col(k))
            if v["comp"] == "is_false":
                df = df.filter(pl.col(k).not_())
        return df


def re_limited(re_meta: pd.DataFrame) -> pd.DataFrame:
    rn = {
        "capacity_mw_nrel_site_lim": "capacity_mw_nrel_site",
        "area_sq_km_lim": "area_sq_km",
        "cum_cap_mw_lim": "cum_cap_mw",
    }
    return (
        re_meta[(re_meta.capacity_mw_nrel_site_lim > 0) & (re_meta.area_sq_km_lim > 0)]
        .drop(columns=rn.values())
        .rename(columns=rn)
        .copy()
    )


def make_core_lhs_rhs(re_meta, limited=False, fossil=True):
    re_meta = re_meta.astype({"icx_genid": str, "re_site_id": str}).copy()
    to_concat = {}
    if fossil:
        to_concat["fossil"] = re_meta.pivot_table(
            index="icx_genid",
            columns="combi_id",
            values="ones",
            aggfunc="first",
            sort=False,
        ).join(
            re_meta.groupby("icx_genid").icx_capacity.first().to_frame("rhs"),
            validate="1:1",
        )
    to_concat["re_area"] = re_meta.pivot_table(
        index="re_site_id",
        columns="combi_id",
        values="area_per_mw",
        aggfunc="first",
        sort=False,
    ).join(
        re_meta.groupby("re_site_id").area_sq_km.max().to_frame("rhs"),
        validate="1:1",
    )
    on_wind = re_meta.query("re_type == 'onshore_wind'")
    to_concat["onshore_wind"] = on_wind.pivot_table(
        index="re_site_id", columns="combi_id", values="ones", aggfunc="first"
    ).join(
        on_wind.groupby("re_site_id").capacity_mw_nrel_site.max().to_frame("rhs"),
        how="left",
        validate="1:1",
    )

    off_wind = re_meta.query("re_type == 'offshore_wind'")
    to_concat["offshore_wind"] = off_wind.pivot_table(
        index="re_site_id", columns="combi_id", values="ones", aggfunc="first"
    ).join(
        off_wind.groupby("re_site_id").capacity_mw_nrel_site.max().to_frame("rhs"),
        how="left",
        validate="1:1",
    )
    if limited:
        if fossil:
            to_concat["fossil"] = to_concat["fossil"].join(
                re_meta.groupby("icx_genid").icx_capacity.first().to_frame("rhs_lim"),
                validate="1:1",
            )
        to_concat["re_area"] = to_concat["re_area"].join(
            re_meta.groupby("re_site_id").area_sq_km_lim.max().to_frame("rhs_lim"),
            validate="1:1",
        )
        to_concat["onshore_wind"] = to_concat["onshore_wind"].join(
            on_wind.groupby("re_site_id").capacity_mw_nrel_site_lim.max().to_frame("rhs_lim"),
            validate="1:1",
        )
        to_concat["offshore_wind"] = to_concat["offshore_wind"].join(
            off_wind.groupby("re_site_id").capacity_mw_nrel_site_lim.max().to_frame("rhs_lim"),
            validate="1:1",
        )
    lhs_rhs_core = pd.concat(to_concat, names=["cons_set", "res_id"]).fillna(0.0)[
        to_concat["re_area"].columns
    ]

    return lhs_rhs_core


def lstrip0(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """Remove leading zeros from column"""
    df[col] = df[col].astype(str).str.lstrip("0")
    return df


def seconds_since_touch(file: Path) -> float:
    if file.exists():
        return time.time() - file.stat().st_mtime
    return time.time()


def fix_cems_datetime(
    df: pd.DataFrame,
    date_col: str = "OP_DATE",
    hour_col: str = "OP_HOUR",
    datetime_col: str = "datetime",
) -> pd.DataFrame:
    df[datetime_col] = pd.to_datetime(df.pop(date_col).astype("object")) + pd.to_timedelta(
        df.pop(hour_col), unit="h"
    )
    return df


def combine_cols(
    df: pd.DataFrame, col1: str, col2: str, out_col: str, sep: str = "_", drop=False
) -> pd.DataFrame:
    df[out_col] = df[col1].astype(str).str.cat(df[col2].astype(str), sep=sep)
    if drop:
        return df.drop([col1, col2], axis=1)
    return df


def check_lat_lons(df: pd.DataFrame) -> pd.DataFrame:
    """Add column that says if lat/lon is safe based on state bounds"""
    df = df.merge(
        pd.DataFrame.from_dict(STATE_BOUNDS).reset_index().rename(columns={"index": "state"}),
        on="state",
        how="left",
    )
    df["safe_lat_lon"] = np.where(
        ((np.floor(df.ymin) <= df.latitude) & (df.latitude <= np.ceil(df.ymax)))
        & ((np.floor(df.xmin) <= df.longitude) & (df.longitude <= np.ceil(df.xmax))),
        True,
        False,
    )
    return df


def add_states(df: pd.DataFrame) -> pd.DataFrame:
    assert "plant_id_eia" in df.columns
    states = pd.read_parquet(ROOT_PATH / "patio_data/plant_id_to_state.parquet")
    return df.merge(states, on="plant_id_eia", how="left")


def fix_na_neg_col(df: pd.DataFrame, col: str) -> pd.DataFrame:
    df[col] = np.where(df[col] < 0, 0, df[col].fillna(0))
    return df


def adjust_col_for_pct_owned(
    df: pd.DataFrame, col: str, pct_owned_col: str = "pct_owned"
) -> pd.DataFrame:
    df[col] = df[col] * df[pct_owned_col]
    return df


def cat_multiindex_as_col(df: pd.DataFrame, new_col: str, droplevel=None) -> pd.DataFrame:
    if droplevel:
        df[new_col] = df.index.droplevel(droplevel).map(lambda x: "_".join(map(str, x)))
    else:
        df[new_col] = df.index.map(lambda x: "_".join(map(str, x)))
    return df


def round_coordinates(
    df: pd.DataFrame,
    in_cols: tuple[str, str] = ("latitude", "longitude"),
    out_cols: tuple[str, str] = ("latitude", "longitude"),
    tol: float = 0.5,
) -> pd.DataFrame:
    """Round coordinates to the nearest ``tol`` degree"""
    adj = 1 / tol if tol < 1 else 1
    df[out_cols[0]] = np.around(df[in_cols[0]].astype(float) * adj) / adj
    df[out_cols[1]] = np.around(df[in_cols[1]].astype(float) * adj) / adj
    return df


def as_ufunc(func):
    """Turn ``np.frompyfunc`` into decorator"""
    return np.frompyfunc(func, 2, 1)


@as_ufunc
def isclose(i, j):
    """~equivalent of ``np.isclose`` for arrays with strs"""
    if isinstance(i, str):
        return i == j
    return np.abs(i - j) < 1e04


def show(fig: go.Figure):
    """Display a plotly figure in a notebook where ``fig.show()`` does not work

    :param fig:
    :return:
    """
    from IPython.core.display import Image, display

    display(Image(pio.to_image(fig, format="png")))


def agg_profile(profile: pd.DataFrame | pd.Series, freq="D") -> pd.DataFrame | pd.Series:
    """Adjust frequency of profile to ``freq``

    :param profile: input profile
    :param freq: the time frequency to aggregate to
        "D" for day, "H" for hour
    :return:
    """
    if pd.infer_freq(profile.index) == freq:
        return profile
    return profile.groupby(pd.Grouper(freq=freq)).sum()


def crn(df, l):
    pass


def df_in_half(df):
    """Split model output df in ~half wihtout breaking up plant/prime/fuel"""
    ix = (
        df.iloc[int(len(df) / 2) :]
        .query("pct_of_re == 1 & report_date == '2020-12-01'")
        .index[0]
        + 1
    )
    return df.iloc[:ix, :], df.iloc[ix:, :]


@njit
def distance_arrays(latlon1: np.ndarray, latlon2: np.ndarray) -> np.ndarray:
    return distance(latlon1[:, 0], latlon1[:, 1], latlon2[:, 0], latlon2[:, 1])


@njit
def distance(
    lat1: float | np.ndarray,
    lon1: float | np.ndarray,
    lat2: float | np.ndarray,
    lon2: float | np.ndarray,
):
    """Compute haversine distance between two points"""
    p = 0.017453292519943295
    hav = (
        0.5
        - np.cos((lat2 - lat1) * p) / 2
        + np.cos(lat1 * p) * np.cos(lat2 * p) * (1 - np.cos((lon2 - lon1) * p)) / 2
    )
    return 12742 * np.arcsin(np.sqrt(hav))


# airports = (
#     pl.read_csv("https://ourairports.com/countries/US/airports.csv")
#     .filter(pl.col("type") == "large_airport")
#     .select(pl.col("local_code").alias("airport_code"), "latitude_deg", "longitude_deg")
# )


def pl_distance(df: pl.LazyFrame, lat1, lat2, lon1, lon2) -> pl.LazyFrame:
    p = 0.017453292519943295
    return df.with_columns(
        distance=12742
        * (
            (
                0.5
                - ((pl.col(lat2) - pl.col(lat1)) * p).cos() / 2
                + (pl.col(lat1) * p).cos()
                * (pl.col(lat2) * p).cos()
                * (1 - ((pl.col(lon2) - pl.col(lon1)) * p).cos())
                / 2
            ).sqrt()
        ).arcsin()
    )


def ninja_profile_fix(df, pid, rtype, solar_ilr):
    s = df.set_index(
        pd.date_range(str(df.local_time[0]).rpartition("-")[0], periods=131496, freq="h")
    ).loc["2006-01-01":"2020-12-30", pid]
    if RE_TECH[rtype] == "solar" and solar_ilr > 1.0:
        s = adjust_ilr(s, ilr=solar_ilr)
    s.name = f"{pid}_{RE_TECH[rtype]}"
    return s


def find_na(df):
    if isinstance(df, pd.DataFrame):
        return df[df.isnull().any(axis=1)]  # noqa: PD003
    return df[df.isnull()]  # noqa: PD003


def prep_for_re_ninja(re: pd.DataFrame) -> pd.DataFrame:
    """Set names and years to tell re_ninja what to download"""
    # re["years"] = np.where(re.ninja_code.isna(), "2006-2020", "2020")
    # re["years"] = np.where(re.technology_description == "solar", "2006-2020", re.years)
    re["years"] = "2006-2020"
    # re["num"] = np.where(re.ninja_code.isna(), 15, 1)
    re["num"] = 15
    # re["name"] = np.where(re.ninja_code.isna(), re.plant_id_eia, re.ninja_code)
    print(
        re.groupby(["technology_description", "years"]).agg(
            {"num": np.sum, "plant_id_eia": "count"}
        )
    )
    print(f"{re['num'].sum() / 200 / 24} days")
    return re


def td_to_hours(td: datetime.timedelta) -> int:
    """Convert timedelta to hours"""
    return int(td.days * 24 + td.seconds / 3600)


def mo_to_days(
    monthly: pd.DataFrame | pd.Series, years=(2006, 2020), last_day=30
) -> pd.DataFrame | pd.Series:
    """Take one year of monthly data and repeat it over ``years`` then resample it to
    hourly so every hour has the full month value
    """
    first, last = years
    df = repeat_mo(monthly, years)
    return df.reindex(
        pd.date_range(f"{first}-01-01", f"{last}-12-{last_day}", freq="D"),
        method="ffill",
    )


def repeat_mo(monthly, years=(2006, 2020)):
    """Take one year of monthly data and repeat it over ``years`` then resample it to
    hourly so every hour has the full month value
    """
    assert len(monthly) == 12
    first, last = years
    df = pd.concat([monthly] * (last - first + 1), ignore_index=True)
    df.index = pd.date_range(f"{first}-01-01", f"{last}-12-01", freq="MS")
    return df


def normalize_cems(cems):
    """Normalize cems data monthly for use with monthly 923 data"""
    # aggregate by plant
    cems = cems.fillna(0).groupby(axis=1, level=0).sum()
    # normalize by month
    return cems / cems.groupby(cems.index.month).transform("sum")


def adjust_ilr(
    profile: pd.Series | pd.DataFrame | np.ndarray, ilr: float
) -> pd.Series | pd.DataFrame | np.ndarray:
    return np.minimum(profile * ilr, 1)


def shift_profile_utc(
    profile: pd.Series | pd.DataFrame | np.ndarray,
    new_utc: int,
    old_utc: int,
    logid: str = "",
) -> pd.Series | pd.DataFrame | np.ndarray:
    """Roll profile to shift it from ``old_utc`` offset to ``new_utc`` offset"""
    if new_utc == old_utc:
        return profile
    if isinstance(new_utc, datetime.datetime):
        new_utc = td_to_hours(new_utc.utcoffset())
    elif isinstance(new_utc, datetime.timedelta):
        new_utc = td_to_hours(new_utc)
    LOGGER.info("%s shifting profile utc %g -> %g", logid, old_utc, new_utc)
    if isinstance(profile, pd.Series):
        return pd.Series(
            np.roll(profile, axis=0, shift=(new_utc - old_utc)),
            index=profile.index,
            name=profile.name,
        )
    if isinstance(profile, pd.DataFrame):
        return pd.DataFrame(
            np.roll(profile, axis=0, shift=(new_utc - old_utc)),
            index=profile.index,
            columns=profile.columns,
        )
    return np.roll(profile, axis=0, shift=(new_utc - old_utc))


def pl_df_product[T: pl.LazyFrame | pl.DataFrame](
    left: T,
    right: T,
    suffix: str = "_r",
) -> T:
    """Determine the product (~combinations) of the rows of two dataframes"""
    return left.unique().join(right.unique(), how="cross", suffix=suffix)


def df_product(
    left: pd.DataFrame,
    right: pd.DataFrame,
    left_suffix: str = "_l",
    right_suffix: str = "_r",
) -> pd.DataFrame:
    """Determine the product (~combinations) of the rows of two dataframes"""
    left, right = left.drop_duplicates(), right.drop_duplicates()
    left.columns = [f"{col}{left_suffix}" for col in left.columns]
    right.columns = [f"{col}{right_suffix}" for col in right.columns]
    left_full = left.loc[left.index.repeat(right.shape[0])].reset_index(drop=True)
    right_full = pd.concat([right] * left.shape[0], ignore_index=True)
    return pd.concat(
        [left_full, right_full],
        axis=1,
    )


def combine_profiles(iter_file, solar_ilr):
    """Combine profiles from list of files, fix timezones, remove 12/31/20 and aggregate to daily"""
    errors = []
    profs = []
    for i, file in enumerate(iter_file):
        pid, _, rtype = file.stem.partition("_")
        df = pd.read_parquet(file)
        if len(df) != 131496 or pd.isna(df.local_time[0]):
            errors.append(file)
            continue
        print(i, end=" ", flush=True)
        s = ninja_profile_fix(df, pid, rtype, solar_ilr)
        profs.append(agg_profile(s, freq="D"))
    return pd.concat(profs, axis=1), errors


def idfn(val):
    """Id function for pytest parameterization

    :param val:
    :return:
    """
    if isinstance(val, float):
        return None
    return str(val)


def load_or_rebuild_parquet(
    e_file: Path, p_file: Path, func: Callable[[Path, Path], None]
) -> pd.DataFrame:
    """Load file from ``p_path`` if it is newer than ``e_path`` otherwise run ``func`` to
    create a new parquet at ``p_path`` based on ``e_path`` and return the df
    """
    if p_file.exists() and e_file.lstat().st_mtime <= p_file.lstat().st_mtime:
        return pd.read_parquet(p_file)
    func(e_file, p_file)
    return pd.read_parquet(p_file)


def clean_col(df, col, dtype):
    """Remove blank strings from col and cast to numeric"""
    try:
        df[col] = df[col].astype(str).replace([" ", ""], [np.nan, np.nan]).astype(float)
        if dtype == "Int64":
            df[col] = df[col].astype("Int64")
    except ValueError as exc:
        raise ValueError(f"{col} -> {dtype} {exc!r}") from exc
    return df


def min_i(x0, y0, x1, y1):
    """Max for iterative LPs."""
    return 0.5 * (1 + np.sign(x0 - y0)) * y1 + 0.5 * (1 - np.sign(x0 - y0)) * x1


def _git_commit_info():
    try:
        command = ["git", "log", "-n", "1", "--pretty=tformat:%h %aD", "--date=short"]
        return subprocess.check_output(command, cwd=str(ROOT_PATH)).strip().decode()  # noqa: S603
    except Exception as e:
        LOGGER.info("Git commit info not found %r", e)
        return "no git info found"


def bbb_path():
    bbb_path_ = ROOT_PATH / "BBB Fossil Transition Analysis Inputs.xlsm"
    if not bbb_path_.exists():
        get("patio-restricted/BBB Fossil Transition Analysis Inputs.xlsm", ROOT_PATH)
    return bbb_path_


@lru_cache
def get_year_map(min_year=2021, max_year=2039):
    bbb = read_named_range(bbb_path(), "Fossil_Price_Year_Table").set_index("Year").squeeze()
    LOGGER.warning("Overriding year map 2021 -> 2021, 2022 -> 2022, 2032 -> 2022")
    bbb.loc[2021] = 2021
    bbb.loc[2022] = 2022
    bbb.loc[2032] = 2022
    return bbb.loc[min_year:max_year].to_dict()


def generate_projection_from_historical_pl[T: pl.LazyFrame | pl.DataFrame](
    in_df: T,
    date_col: str = "datetime",
    year_mapper: dict | None = None,
) -> T:
    if year_mapper is None:
        year_mapper = get_year_map(max_year=2039)

    lazy = True
    if isinstance(in_df, pl.DataFrame):
        in_df = in_df.lazy()
        lazy = False

    tz_ = "" if in_df.collect_schema()[date_col].time_zone is None else "%Z"

    out = pl.concat(
        [
            in_df.filter(pl.col(date_col).dt.year() == p_y).with_columns(
                pl.col(date_col)
                .dt.strftime(f"{f_y}-%m-%d %H:%M:%S{tz_}")
                .str.strptime(pl.Datetime, format=f"%Y-%m-%d %H:%M:%S{tz_}")
            )
            for f_y, p_y in year_mapper.items()
        ],
        parallel=True,
    )
    if lazy:
        return out
    return out.collect()


def generate_projection_from_historical(
    in_df: pd.DataFrame,
    date_col: str = "datetime",
    year_mapper: dict | None = None,
) -> pd.DataFrame:
    if in_df.empty:
        return in_df

    if year_mapper is None:
        year_mapper = get_year_map(max_year=2039)
    idx_names = in_df.index.names
    out_dfs = []

    def replacer(s, y1):
        return pd.to_datetime(s.strftime(f"{y1}-%m-%d %H:%M:%S %Z"))

    if (
        date_col in idx_names
        and isinstance(in_df.index, pd.MultiIndex)
        and not isinstance(in_df.columns, pd.MultiIndex)
    ):
        in_df = in_df.reset_index()
        set_index = True
    else:
        set_index = False

    if date_col in in_df:
        for p_yr, h_yr in year_mapper.items():
            temp = in_df[in_df[date_col].dt.year == h_yr].copy()
            temp.loc[:, date_col] = replacer(temp[date_col].dt, p_yr)
            out_dfs.append(temp)
    elif date_col in idx_names and not isinstance(in_df.index, pd.MultiIndex):
        for p_yr, h_yr in year_mapper.items():
            temp = in_df[in_df.index.get_level_values(date_col).year == h_yr]
            temp.index = replacer(temp.index, p_yr)
            out_dfs.append(temp)
    elif isinstance(in_df.columns, pd.MultiIndex):
        raise ValueError(
            f"If `{date_col}` is in a MultiIndex, columns cannot be a Multiindex."
        )
    else:
        raise ValueError(f"`{date_col}` must be a column or index the index of `in_df`")
    out_dfs = pd.concat(out_dfs, axis=0)
    if set_index:
        out_dfs = out_dfs.set_index(idx_names)
    return out_dfs


def read_named_range(path, name):
    with all_logging_disabled():
        from openpyxl.reader.excel import load_workbook

        wb = load_workbook(path)
        ws, reg = list(wb.defined_names[name].destinations)[0]
        ws = wb[ws]
        region = ws[reg]
        df = pd.DataFrame(
            ([cell.value for cell in row] for row in region[1:]),
            columns=[cell.value for cell in region[0]],
        )
    return df


def solver():
    return os.environ.get("SOLVER", cp.COPT)
